"""Module book.py"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from logistiki.utils import dec, dec2gr, startswith_any, dec2grp, date_iso2gr
from logistiki.utils import levels_reverse
from logistiki.logger import logger


def trimino(isodate: str) -> str:
    """Returns Etos and trimino ('2021-09-17' -> '20213')"""
    yyy, mmm, _ = isodate.split("-")
    trim = {
        "01": "1",
        "02": "1",
        "03": "1",
        "04": "2",
        "05": "2",
        "06": "2",
        "07": "3",
        "08": "3",
        "09": "3",
        "10": "4",
        "11": "4",
        "12": "4",
    }
    return f"{yyy}{trim[mmm]}"


# def date_iso2gr(iso_date: str) -> str:
#     """YYYY-MM-DD -> DD/MM/YYYY"""
#     if iso_date == "" or iso_date is None:
#         return ""
#     yyyy, mmm, ddd = iso_date.split("-")
#     return f"{ddd}/{mmm}/{yyyy}"


def clean_per(per: str) -> str:
    """Καθάρισμα περιγραφών"""
    per = per.replace("-Τιμολόγιο Αγορών (Προμηθευτή)", "")
    per = per.replace("-Τιμολόγιο Αγορών (Προμηθευτή", "")
    per = per.replace("-Τιμολόγιο Πώλησης", "")
    per = per.replace("- Δ.Α. Αγορών (Προμηθευτή)", "")
    per = per.replace("-Πιστωτικό Τιμολόγιο Επιστροφής", "")
    per = per.replace("-Δ.Αποστολής - Τιμολόγιο Πώλησης", "")
    per = per.replace("-Δ.Αποστολής - Τιμολόγιο Πώλησ", "")
    per = per.replace("-Πιστωτικό Τιμολόγιο - Δ.Α. Αγορών (Προμηθευ", "")
    per = per.replace("-Δ.Αποστολής - Τιμολόγιο Πώ", "")
    return per


def ee_value_sign(acline):
    """Αλλαγή προσήμου ανάλογα με το αν είναι Χρέωση ή Πίστωση"""
    if acline["account"][0] in "126":
        if acline["typ"] == 1:
            return acline["value"]
        else:
            return -acline["value"]
    elif acline["account"][0] == "7":
        if acline["typ"] == 1:
            return -acline["value"]
        else:
            return acline["value"]
    else:
        return acline["value"]


def closest_name(code: str, chart: dict) -> str:
    """
    Get closest possible name for the account
    """
    if not chart:
        return ""
    levels = levels_reverse(code)
    for level in levels:
        if level in chart:
            return chart[level]
    return ""


class Book:
    """Double entry accounting book"""

    def __init__(self, co_data, transactions, accounts, afms) -> None:
        self.name = co_data["name"]
        self.afm = co_data["afm"]
        self.year = co_data["year"]
        self.branch = co_data["branch"]
        self.accounts = accounts  # Dictionary with accounts : names
        self.transactions = transactions  # list of dict of transactions
        self.afms = afms

    def trans_print(self, tran_id: int) -> None:
        """Print a Transaction on terminal"""
        trans = self.transactions[tran_id - 1]
        stt = f"┌─◆ Άρθρο {trans['id']}\n"
        stt += "│\n"
        stt += f"│ Ημερομηνία  : {date_iso2gr(trans['date'])}\n"
        stt += f"│ Παραστατικό : {trans['partype']} {trans['parno']}\n"
        stt += f"│ Περιγραφή   : {trans['perigrafi']} {trans['perigr2']}\n"
        stt += f"│ ΑΦΜ         : {trans['afm']}\n"
        stt += "│\n"
        ltm = "│ {acc:<13} {acp:<40} {deb:>14} {cre:>14} \n"
        stt += ltm.format(
            acc="Λογαριασμός", acp="Περιγραφή", deb="Χρέωση", cre="Πίστωση"
        )
        stt += ltm.format(
            acc="-----------", acp="---------", deb="------", cre="-------"
        )
        # stt += '│ ' + '-' * 50 + '\n'
        # stt += ltm.format(acc='-----------', deb='------', cre='-------')
        for lin in trans["lines"]:
            deb = dec2gr(lin["value"] if lin["typ"] == 1 else 0)
            cre = dec2gr(lin["value"] if lin["typ"] == 2 else 0)
            acp = self.accounts[lin["account"]]
            stt += ltm.format(acc=lin["account"], acp=acp, deb=deb, cre=cre)
        stt += "└─►\n"
        print(stt)

    def isozygio(self, apo=None, eos=None, chart=None) -> str:
        """
        Ισοζύγιο λογαριασμών
        """
        iso = {}
        txr = tpi = typ = 0
        for tra in self.filter_by_dates(apo, eos):
            for lin in tra["lines"]:
                debit = lin["value"] if lin["typ"] == 1 else 0
                credit = lin["value"] if lin["typ"] == 2 else 0
                gcode = lin["account"]
                iso[gcode] = iso.get(
                    gcode, {"debit": dec(0), "credit": dec(0)})
                iso[gcode]["debit"] += debit
                iso[gcode]["credit"] += credit
                txr += debit
                tpi += credit
                typ += debit - credit
        fiso = {}
        apot = 0
        for code, val in iso.items():
            for lcode in levels_reverse(code):
                fiso[lcode] = fiso.get(
                    lcode, {"debit": dec(0), "credit": dec(0)})
                fiso[lcode]["debit"] += val["debit"]
                fiso[lcode]["credit"] += val["credit"]
            if code[0] in "267":
                apot += val["debit"] - val["credit"]
        tapo = f" Από: {date_iso2gr(apo)}"
        teos = f" Έως: {date_iso2gr(eos)}"
        tapoeos = f"{tapo}{teos}"
        tmp = f"\n{'ΙΣΟΖΥΓΙΟ ΛΟΓΙΣΤΙΚΗΣ':^104}\n"
        tmp += f"{tapoeos:^104}\n"
        tmp += "-" * 104 + "\n"
        for acc in sorted(fiso):
            delta = dec2gr(fiso[acc]["debit"] - fiso[acc]["credit"])
            if acc in self.accounts:
                accp = self.accounts[acc]
            else:
                accp = closest_name(acc, chart)
            tmp += (
                f"{acc:<15}{accp:<50} {dec2gr(fiso[acc]['debit']):>12} "
                f"{dec2gr(fiso[acc]['credit']):>12} {delta:>12}\n"
            )
        tmp += "-" * 104 + "\n"
        tmp += (
            f"{'':<15}{'Σύνολα':^50} {dec2gr(txr):>12} "
            f"{dec2gr(tpi):>12} {dec2gr(typ):>12}\n"
            f"{'':<15}{'Αποτέλεσμα (2, 6, 7)':^50} {dec2gr(0):>12} "
            f"{dec2gr(0):>12} {dec2gr(apot):>12}\n"
        )
        return tmp

    def filter_by_dates(self, apo: str, eos: str):
        """Filtering Transactions by dates"""
        for transaction in self.transactions:
            if apo and transaction["date"] < apo:
                continue
            if eos and transaction["date"] > eos:
                continue
            yield transaction

    def totals_for_fpa(self, apo, eos):
        """
        Ισοζύγιο ειδικά για τον υπολογισμό του ΦΠΑ
        """
        iso = {}
        fpa_delta = 0
        for tra in self.filter_by_dates(apo, eos):
            for lin in tra["lines"]:
                account = lin["account"]
                if account.startswith("54.00"):
                    if not account.startswith("54.00.99"):
                        fpa_delta += lin["value"] if lin["typ"] == 1 else - \
                            lin["value"]
                if account[0] not in "1267":
                    continue
                iso[account] = iso.get(account, dec(0))
                if account[0] in "126":
                    iso[account] += lin["value"] if lin["typ"] == 1 else -lin["value"]
                elif account[0] == "7":
                    iso[account] += lin["value"] if lin["typ"] == 2 else -lin["value"]
        iso[5400] = fpa_delta
        return iso

    def fpa(self, apo=None, eos=None):
        """
        Αναφορά ΦΠΑ για την περίοδο apo-eos
        """
        return f"fpa (apo={apo}, eos={eos})Not Implemented"

    def ee_book(self) -> list:
        """
        Βιβλίο εσόδων-εξόδων
        """
        eebook = []
        counter = 0
        for trn in self.transactions:
            if not trn["is_ee"]:
                continue
            counter += 1
            tval = tfpa = 0
            typos = set()
            accounts = set()
            for lin in trn["lines"]:
                if lin["account"][0] in "267":
                    accounts.add(lin["account"])
                    tval += ee_value_sign(lin)
                    typos.add(lin["account"][0])
                elif lin["account"].startswith("54.00."):
                    tfpa += ee_value_sign(lin)
            eebook.append(
                {
                    "typos": "".join(sorted(list(typos))),
                    "id": counter,
                    "date": trn["date"],
                    "part": trn["partype"],
                    "parno": trn["parno"],
                    "afm": trn["afm"],
                    "perigrafi": trn["perigrafi"],
                    "value": tval,
                    "fpa": tfpa,
                    "total": tval + tfpa,
                    "accounts": accounts,
                }
            )
        return eebook

    def ee_book_report(self, exclude=None, only=None):
        """
        Creates an ee report
        exclude: accounts to exclude from report
        only: only theese accounts to include to report
        """
        tot2 = tot6 = tot7 = 0
        fpa2 = fpa6 = fpa7 = 0
        for trn in self.ee_book():
            if only:
                if not startswith_any(trn["accounts"], only):
                    continue
            if exclude:
                if startswith_any(tuple(trn["accounts"]), exclude):
                    continue
            # Εδώ κάνουμε τις αθροίσεις
            if trn["typos"] == "2":
                tot2 += trn["value"]
                fpa2 += trn["fpa"]
            if trn["typos"] == "6":
                tot6 += trn["value"]
                fpa6 += trn["fpa"]
            if trn["typos"] == "7":
                tot7 += trn["value"]
                fpa7 += trn["fpa"]
            trn["eper"] = trn["perigrafi"][:50]
            print(
                "{id:<5}{typos:<3}{date} {afm:<9} "
                "{parno:<20} {eper:<50}{value:>12}"
                "{fpa:>12}{total:>12} {accounts}".format(**trn)
            )
        print("")
        print("=" * 40)
        print(f"  Ομάδες       {'Αξία':^12} {'ΦΠΑ':^12}")
        print("-" * 40)
        print(f"Ομάδα 2      : {tot2:>12} {fpa2:>12}")
        print(f"Ομάδα 6      : {tot6:>12} {fpa6:>12}")
        print(f"Σύνολο 2+6   : {tot6+tot2:>12} {fpa6+fpa2:>12}")
        print(f"Ομάδα 7      : {tot7:>12} {fpa7:>12}")
        print(f"Κέρδος 7-2-6 : {tot7-tot6-tot2:>12} {fpa7-fpa6-fpa2:>12}")
        print("=" * 40)

    def _ee_bookd(self, el2ee):
        """
        Βιβλίο εσόδων-εξόδων για excell
        """
        eebook = []
        counter = 0
        for trn in self.transactions:
            if trn["is_ee"]:
                counter += 1
                tts = {}
                typos = set()
                accounts = set()
                for lin in trn["lines"]:
                    if lin["account"][0] in "267":
                        accounts.add(lin["account"])
                        typos.add(lin["account"][0])
                        ee_key = el2ee[lin["account"]]
                        tts[ee_key] = tts.get(ee_key, 0) + ee_value_sign(lin)
                    elif lin["account"].startswith("54.00."):
                        ee_key = el2ee[lin["account"]]
                        tts[ee_key] = tts.get(ee_key, 0) + ee_value_sign(lin)
                tts["typos"] = "".join(sorted(list(typos)))
                tts["id"] = counter
                tts["date"] = date_iso2gr(trn["date"])
                tts["trimino"] = trimino(trn["date"])
                tts["part"] = trn["partype"]
                tts["par"] = trn["parno"]
                tts["afm"] = trn["afm"]
                tts["per"] = clean_per(trn["perigrafi"])
                tts["accounts"] = accounts
                eebook.append(tts)
        return eebook

    def ee_book2excel(self, cfg, filename="txt.xlsx"):
        """Βιβλίο Εσόδων-Εξόδων σε excell αρχείο"""
        el2ee = dict(cfg["eeaccounts"])
        eecols = dict(cfg["eecols"])
        # Έλεγχος εάν όλες οι τιμές του el2ee είναι κλειδιά στον eecols
        for val in el2ee.values():
            if val not in eecols.keys():
                raise ValueError(f"Η στήλη {val} δεν έχει περιγραφή.")
        wb1 = Workbook()
        # date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
        sheet = wb1.active
        sheet.page_setup.orientation = 'landscape'  # sheet.ORIENTATION_LANDSCAPE
        sheet.row_dimensions[1].height = 70
        sheet.row_dimensions[1].font = Font(name="Liberation Sans", bold=True)
        sheet.row_dimensions[1].alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )

        for i, val in enumerate(eecols.values()):
            sheet.cell(column=i + 1, row=1, value=val)

        for i, trn in enumerate(self._ee_bookd(el2ee)):
            # print(trn)
            for j, ee_key in enumerate(eecols):
                # print(trn[ee_key])
                trn.get(ee_key, "")
                sheet.cell(column=j + 1, row=i + 2, value=trn.get(ee_key, ""))

        wb1.save(filename)

    def myf_xml(self, exclude=None, only=None, koybas=(), rfpa=(), action="replace"):
        """
        Συγκεντρωτικές τιμολογίων πελατών/προμηθευτών σε xml
        """
        data = {
            "action": action,
            "co": {
                "afm": self.afm,
                "month": "12",
                "year": self.year,
                "branch": self.branch,
            },
            "gcash": [],
            "oexpenses": {},
        }
        dat = "2020-12-31"
        # di1 = {}
        di2 = {}
        di6 = {}
        di7 = {}
        cre = "credit"
        nor = "normal"
        tam = {"tamNo": "", "amount": 0, "tax": 0, "date": dat}
        reversed_afm = set()
        for trn in self.ee_book():
            # first we filter records
            if only:
                if not startswith_any(trn["accounts"], only):
                    continue
            if exclude:
                if startswith_any(tuple(trn["accounts"]), exclude):
                    continue

            afm = trn["afm"]

            if trn["typos"] == "1":
                pass

            elif trn["typos"] in "26":
                # Εάν έχουμε λογαριασμούς χωρίς έκπτωση φπα, επειδή οι
                # εγγραφές είναι συνολικές θα πρέπει πρωτού προχωρήσουμε
                # να αποφορολογήσουμε
                if any(i in rfpa for i in trn["accounts"]):
                    if trn["fpa"] == 0:
                        val = dec(trn["value"] / dec(1.24))
                        trn["fpa"] = trn["value"] - val
                        trn["value"] = val
                        reversed_afm.add(afm)

                # Προμηθευτές σε κουβά (πχ ΔΕΗ)
                if afm in koybas:
                    data["oexpenses"]["amount"] = data["oexpenses"].get(
                        "amount", 0)
                    data["oexpenses"]["amount"] += trn["value"]
                    data["oexpenses"]["tax"] = data["oexpenses"].get("tax", 0)
                    data["oexpenses"]["tax"] += trn["fpa"]
                    data["oexpenses"]["date"] = data["oexpenses"].get(
                        "date", dat)
                    continue
                if len(afm) < 9:
                    raise ValueError(f"There is an error in {trn}")
                # Διαφορετικά κανονικά ανά ΑΦΜ/normal-credit
                di6[afm] = di6.get(afm, {})
                if trn["value"] >= 0:
                    di6[afm][nor] = di6[afm].get(
                        nor,
                        {
                            "afm": afm,
                            "amount": 0,
                            "tax": 0,
                            "note": nor,
                            "invoices": 0,
                            "date": dat,
                            "nonObl": 0,
                        },
                    )
                    di6[afm][nor]["amount"] += trn["value"]
                    di6[afm][nor]["tax"] += trn["fpa"]
                    di6[afm][nor]["invoices"] += 1
                else:
                    di6[afm][cre] = di6[afm].get(
                        cre,
                        {
                            "afm": afm,
                            "amount": 0,
                            "tax": 0,
                            "note": cre,
                            "invoices": 0,
                            "date": dat,
                            "nonObl": 0,
                        },
                    )
                    di6[afm][cre]["amount"] -= trn["value"]
                    di6[afm][cre]["tax"] -= trn["fpa"]
                    di6[afm][cre]["invoices"] += 1

            elif trn["typos"] == "7":
                if afm == "":
                    tam["amount"] += trn["value"]
                    tam["tax"] += trn["fpa"]
                else:
                    di7[afm] = di7.get(afm, {})
                    if trn["value"] >= 0:
                        di7[afm][nor] = di7[afm].get(
                            nor,
                            {
                                "afm": afm,
                                "amount": 0,
                                "tax": 0,
                                "note": nor,
                                "invoices": 0,
                                "date": dat,
                            },
                        )
                        di7[afm][nor]["amount"] += trn["value"]
                        di7[afm][nor]["tax"] += trn["fpa"]
                        di7[afm][nor]["invoices"] += 1
                    else:
                        di7[afm][cre] = di7[afm].get(
                            cre,
                            {
                                "afm": afm,
                                "amount": 0,
                                "tax": 0,
                                "note": cre,
                                "invoices": 0,
                                "date": dat,
                            },
                        )
                        di7[afm][cre]["amount"] -= trn["value"]
                        di7[afm][cre]["tax"] -= trn["fpa"]
                        di7[afm][cre]["invoices"] += 1

            else:
                raise ValueError("Αδύνατη περίπτωση !!!")
        logger.info(f"Λογαριασμοί που εξαιρούνται σπό τη ΜΥΦ: {exclude}")
        logger.info(f"ΜΥΦ μόνο για λογαριασμούς: {only}")
        logger.info(f"ΑΦΜ για κουβά εξόδων: {koybas}")
        logger.info(f"Λογαριασμοί για αντιστροφή ΦΠΑ: {rfpa}")
        logger.info(f"ΑΦΜ που έγινε αντιστροφή ΦΠΑ: {reversed_afm}")
        # Έχουμε πλέον τα σύνολα ανα ομάδα
        gexpenses = []
        for afm, decre in sorted(di2.items()):
            for _, val in decre.items():
                val["amount"] = dec2grp(val["amount"])
                val["tax"] = dec2grp(val["tax"])
                gexpenses.append(val)
        # ΕΛΛΗΝΙΚΟ format ΑΡΙΘΜΩΝ
        # Εγγραφές χωρίς έκπτωση του ΦΠΑ επαναφορά
        for afm, decre in sorted(di6.items()):
            for _, val in decre.items():
                val["amount"] = dec2grp(val["amount"])
                val["tax"] = dec2grp(val["tax"])
                gexpenses.append(val)
        data["gexpenses"] = gexpenses

        grevenues = []
        for afm, decre in sorted(di7.items()):
            for _, val in decre.items():
                val["amount"] = dec2grp(val["amount"])
                val["tax"] = dec2grp(val["tax"])
                grevenues.append(val)
        data["grevenues"] = grevenues
        # Mόνο όταν υπάρχουν τιμές στο gcash
        if tam["amount"] != 0:
            tam["amount"] = dec2grp(tam["amount"])
            tam["tax"] = dec2grp(tam["tax"])
            data["gcash"].append(tam)
        # Μόνο όταν υπάρχουν τιμές στο oexpenses
        if data["oexpenses"]:
            data["oexpenses"]["amount"] = dec2grp(data["oexpenses"]["amount"])
            data["oexpenses"]["tax"] = dec2grp(data["oexpenses"]["tax"])
        return data

    def kartella(self, account, apo=None, eos=None):
        """
        Μας δίνει την καρτέλλα του λογαριασμού
        id ημερομηνία παραστατικό περιγραφή χρεωση πίστωση υπόλοιπο
        -------------------------------------------------------------
                      από μεταφορά            100.34  30.88     20.33
        12  10/1/2020   ΤΔΑ322      αγορες     10.35
        """
        dapo = apo if apo else "1900-01-01"
        deos = eos if eos else "3000-12-31"
        if dapo > deos:
            deos = dapo
        tdelta = dec(0)
        prin = {
            "id": "",
            "dat": "",
            "acc": "",
            "par": "Aπό μεταφορά",
            "per": "",
            "debit": dec(0),
            "credit": dec(0),
            "delta": dec(0),
        }
        per = []
        meta = {
            "id": "",
            "dat": "",
            "acc": "",
            "par": "Επόμενες εγγραφές",
            "per": "",
            "debit": dec(0),
            "credit": dec(0),
            "delta": dec(0),
        }
        per_debit = per_credit = 0
        for trn in self.transactions:
            for lin in trn["lines"]:
                debit = lin["value"] if lin["typ"] == 1 else 0
                credit = lin["value"] if lin["typ"] == 2 else 0
                if lin["account"].startswith(account):
                    if trn["date"] < dapo:
                        prin["debit"] += debit
                        prin["credit"] += credit
                        tdelta += debit - credit
                        prin["delta"] = tdelta
                    elif dapo <= trn["date"] <= deos:
                        per_debit += debit
                        per_credit += credit
                        tdelta += debit - credit
                        adi = {
                            "id": trn["id"],
                            "dat": trn["date"],
                            "acc": lin["account"],
                            "par": trn["parno"][:20],
                            "per": trn["perigrafi"][:45],
                            "debit": dec2gr(debit),
                            "credit": dec2gr(credit),
                            "delta": dec2gr(tdelta),
                        }
                        per.append(adi)
                    else:
                        meta["debit"] += debit
                        meta["credit"] += credit
                        tdelta += debit - credit
                        meta["delta"] = tdelta
        if prin["delta"] == 0:
            dprin = []
        else:
            prin["debit"] = dec2gr(prin["debit"])
            prin["credit"] = dec2gr(prin["credit"])
            prin["delta"] = dec2gr(prin["delta"])
            dprin = [prin]

        data = dprin + per
        tit = {
            "id": "AA",
            "dat": "Ημ/νία",
            "acc": "Λογαριασμός",
            "par": "Παραστατικό",
            "per": "Περιγραφή",
            "debit": "Χρέωση",
            "credit": "Πίστωση",
            "delta": "Υπόλοιπο",
        }
        fst = (
            "{id:<4} {dat:10} {acc:<15} {par:<20} {per:<45} "
            "{debit:>12} {credit:>12} {delta:>12}\n"
        )
        st1 = f"\nΚαρτέλλα λογαριασμού {account}\n"
        if apo:
            st1 += f"Aπό: {date_iso2gr(dapo)}\n"
        if eos:
            st1 += f"Έως: {date_iso2gr(deos)}\n"
        st1 += fst.format(**tit)
        st1 += "-" * 137 + "\n"
        for line in data:
            st1 += fst.format(**line)
        synola_periodoy = {
            "id": "",
            "dat": "",
            "acc": "",
            "par": "Σύνολα περιόδου",
            "per": "",
            "debit": dec2gr(per_debit),
            "credit": dec2gr(per_credit),
            "delta": dec2gr(0),
        }
        st1 += fst.format(**synola_periodoy)
        if meta["delta"] != 0:
            meta["debit"] = dec2gr(meta["debit"])
            meta["credit"] = dec2gr(meta["credit"])
            meta["delta"] = dec2gr(meta["delta"])
            st1 += fst.format(**meta)
        return st1

    def __repr__(self):
        return (
            f"Book(name='{self.name}', "
            f"afm={self.afm}, "
            f"year={self.year}, "
            f"branch='{self.branch}', "
            f"transactions={self.transaction_number}, "
            f"accounts={self.account_number})"
        )

    @property
    def account_number(self):
        """
        Αριθμός λογαριασμών
        """
        return len(self.accounts)

    @property
    def transaction_number(self):
        """
        Αριθμός άρθρων
        """
        return len(self.transactions)
